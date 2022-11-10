import pandas as pd
import numpy as np
import pathlib
from ..helpers import date_helpers
from . import dataframe
import shutil
import openpyxl
import glob, os


# Bids
def read_bid(campaign_type, path_bid, filename='bulk.xlsx'):
    if campaign_type == 'SP':
        ws_name = 'Sponsored Products Campaigns'
    if campaign_type == 'SB' or campaign_type == 'SBV':
        ws_name = 'Sponsored Brands Campaigns'
    if campaign_type == 'SD':
        ws_name = 'Sponsored Display Campaigns'
    path = pathlib.Path(path_bid + '/raw/' + filename)

    if campaign_type == 'SP':
        df_bid = pd.read_excel(io=path, sheet_name=ws_name, header=0, engine="openpyxl",
                               converters={'Campaign Id': str, 'Ad Group Id': str, 'Keyword Id (Read only)': str,
                                           'Product Targeting Id (Read only)': str})  # index_col=0)
        df_bid['Campaign Type'] = campaign_type
        df_bid = dataframe.get_cpc(df_bid)
        df_bid['Targeting'] = df_bid.apply(lambda x: x['Keyword Text'] if x['Entity'] == 'Keyword' else x[
            'Resolved Product Targeting Expression (Informational only)'], axis=1)
        df_bid['Keyword Id (Read only)'] = df_bid['Keyword Id (Read only)'].fillna(0)  # .astype('Int64').astype('str')
        df_bid['Product Targeting Id (Read only)'] = df_bid['Product Targeting Id (Read only)'].fillna(
            0)  # .astype('Int64').astype('str')
        df_bid['Campaign Name'] = df_bid['Campaign Name (Informational only)']
        df_bid['Ad Group Name'] = df_bid['Ad Group Name (Informational only)']
        df_bid['Target Id'] = df_bid.apply(lambda x: x['Keyword Id (Read only)'] if x['Entity'] == 'Keyword' else x[
            'Product Targeting Id (Read only)'], axis=1)

    if campaign_type == 'SB' or campaign_type == 'SBV':
        df_bid = pd.read_excel(io=path, sheet_name=ws_name, header=0, engine="openpyxl",
                               converters={'Campaign Id': str, 'Ad Group Id (Read only)': str,
                                           'Keyword Id (Read only)': str,
                                           'Product Targeting Id (Read only)': str})  # index_col=0)
        df_bid['Campaign Type'] = campaign_type
        df_bid = dataframe.get_cpc(df_bid)
        df_bid['Targeting'] = df_bid.apply(
            lambda x: x['Keyword Text'] if x['Entity'] == 'Keyword' else x['Product Targeting Expression'], axis=1)
        df_bid['Keyword Id (Read only)'] = df_bid['Keyword Id (Read only)'].fillna(0)  # .astype('Int64').astype('str')
        df_bid['Product Targeting Id (Read only)'] = df_bid['Product Targeting Id (Read only)'].fillna(
            0)  # .astype('Int64').astype('str')
        df_bid['Campaign Name'] = df_bid['Campaign Name (Informational only)']
        df_bid['Ad Group Name'] = 'Not available'
        df_bid['Ad Group Id'] = df_bid['Ad Group Id (Read only)'].fillna(0).astype(str)
        df_bid['Target Id'] = df_bid.apply(lambda x: x['Keyword Id (Read only)'] if x['Entity'] == 'Keyword' else x[
            'Product Targeting Id (Read only)'], axis=1)

    if campaign_type == 'SD':
        df_bid = pd.read_excel(io=path, sheet_name=ws_name, header=0, engine="openpyxl",
                               converters={'Campaign Id': str, 'Ad Group Id': str,
                                           'Targeting Id (Read only)': str})  # index_col=0)
        df_bid['Campaign Type'] = campaign_type
        df_bid = dataframe.get_cpc(df_bid)

        df_bid['Targeting'] = df_bid['Resolved Targeting Expression (Informational only)']
        # df_bid['Keyword Text']=df_bid.apply(lambda x: x['Keyword Text'] if x['Entity']=='Keyword' else  x['Resolved Targeting Expression (Informational only)'], axis=1)
        df_bid['Product Targeting Expression'] = df_bid['Targeting Expression'].fillna(0).astype(str)
        # df_bid['Targeting']=df_bid['Product Targeting Expression']
        df_bid['Product Targeting Id (Read only)'] = '\'' + df_bid['Targeting Id (Read only)'].fillna(0).astype(str)
        df_bid['Campaign Name'] = df_bid['Campaign Name (Informational only)']
        df_bid['Ad Group Name'] = df_bid['Ad Group Name (Informational only)']
        df_bid['Match Type'] = "-"
        df_bid['Target Id'] = df_bid['Targeting Id (Read only)']

    df_bid['Campaign Id'] = '\'' + df_bid['Campaign Id'].fillna(0).astype(str)
    df_bid['Ad Group Id'] = '\'' + df_bid['Ad Group Id'].fillna(0).astype(str)
    df_bid['Target Id'] = '\'' + df_bid['Target Id'].fillna(0).astype(str)  # .astype('Int64').astype('str')
    df_bid['Match Type'] = df_bid.apply(lambda x: '-' if x['Entity'] == 'Product Targeting' else x['Match Type'],
                                        axis=1)
    return df_bid


def append_bid_history_with_date(path, campaign_types=['SP', 'SB', 'SD']):
    all_files = glob.glob(path + "/raw/bulk*.xlsx")
    for file in all_files:
        filename = os.path.basename(file)
        print(filename)
        for campaign_type in campaign_types:

            df_bid = read_bid(campaign_type, path, filename)
            creation_date = date_helpers.file_date(path + '/raw/' + filename)
            df_bid = dataframe.add_date_dataframe(df_bid, creation_date)
            df_bid['Campaign Type'] = campaign_type
            df_bid = dataframe.to_datetime64(df_bid)
            try:
                df_bid_history = read_bid_history(campaign_type, path)
            except:
                df_bid_history = df_bid
            df_bid_history = dataframe.to_datetime64(df_bid_history)
            df_bid_history = pd.concat([df_bid_history, df_bid])  # .reset_index(drop=True)
            df_bid_history.drop_duplicates(inplace=True, ignore_index=True)
            write_bid_history(df_bid_history, campaign_type, path)


def read_bid_history(campaign_type, path):
    df = pd.read_csv(path + "/prediction/bid_history_" + campaign_type + ".csv")
    if campaign_type == 'SP':
        df['Bid'] = df.apply(lambda x: x['Bid'] if x['Bid'] > 0 else x['Ad Group Default Bid (Informational only)'],
                             axis=1)

    if campaign_type == 'SB':
        df['Ad Group Name'] = 'Not available'
    if campaign_type == 'SD':
        df['Match Type'] = "-"
        df['Bid'] = df.apply(lambda x: x['Ad Group Default Bid (Informational only)'] if x['Bid'] == '' else x['Bid'],
                             axis=1)

    df['Campaign Id'] = df['Campaign Id'].fillna(0).astype(str)
    df['Ad Group Id'] = df['Ad Group Id'].fillna(0).astype(str)
    df['Target Id'] = df['Target Id'].fillna(0).astype(str)
    # df=dataframe.select_row_by_val(df,'State', 'enabled' )

    return df


class NotFound(Exception):
    def __init__(self, value):
        self.value = value

    def __str__(self):
        return repr(self.value)


def write_bid_history(dfbid_history, campaign_type, path):
    dfbid_history.to_csv(path + "/prediction/bid_history_" + campaign_type + ".csv", index=False)


# Performance report
def read_report(campaign_type, path):
    if campaign_type == 'SP':
        sheet_name = 'Sponsored Product Keyword Repor'
        path_file = path + '/raw/Sponsored Products Targeting report.xlsx'
    if campaign_type == 'SB':
        sheet_name = 'Sponsored Brands Keyword Report'
        path_file = path + '/raw/Sponsored Brands Keyword report.xlsx'
    if campaign_type == 'SBV':
        sheet_name = 'Sponsored Brands + Video Keywor'
        path_file = path + '/raw/Sponsored Brands Video Keyword report.xlsx'
    if campaign_type == 'SD':
        sheet_name = 'Sponsored Display Targeting Rep'
        path_file = path + '/raw/Sponsored Display Targeting report.xlsx'

    df = pd.read_excel(io=path_file, sheet_name=sheet_name, header=0, engine="openpyxl")
    # df['Date']=pd.to_datetime(df['Date']).dt.date
    df['Date'] = pd.to_datetime(df['Date'], unit='d')

    if campaign_type == 'SP' or campaign_type == 'SB' or campaign_type == 'SBV':
        df['Match Type'] = df['Match Type'].replace({'EXACT': 'exact', 'PHRASE': 'phrase', 'BROAD': 'broad'})
        df = dataframe.change_val_if_col_contains(df, 'Targeting', 'asin', 'Match Type', 'Targeting Expression')
        df = dataframe.change_val_if_col_contains(df, 'Targeting', 'category', 'Match Type', 'Targeting Expression')
        df = dataframe.change_val_if_col_contains(df, 'Targeting', 'loose-match', 'Match Type',
                                                  'Targeting Expression Predefined')
        df = dataframe.change_val_if_col_contains(df, 'Targeting', 'complements', 'Match Type',
                                                  'Targeting Expression Predefined')
        df = dataframe.change_val_if_col_contains(df, 'Targeting', 'substitutes', 'Match Type',
                                                  'Targeting Expression Predefined')
        df = dataframe.change_val_if_col_contains(df, 'Targeting', 'close-match', 'Match Type',
                                                  'Targeting Expression Predefined')
    if campaign_type == 'SP':
        try:
            df['Conversions'] = df['7 Day Total Orders (#)']  # 7 Day Total Units (#)']
            try:
                df['Sales'] = df['7 Day Total Sales ']  # US
            except KeyError:
                df['Sales'] = df['7 Day Total Sales']  # EU
        except KeyError:
            df['Conversions'] = df['14 Day Total Orders (#)']  # Vendor
            df['Sales'] = df['14 Day Total Sales ']

    if campaign_type == 'SB' or campaign_type == 'SBV':
        df['Conversions'] = df['14 Day Total Orders (#)']  # 14 Day Total Units (#)']
        try:
            df['Sales'] = df['14 Day Total Sales ']  # US
        except:
            df['Sales'] = df['14 Day Total Sales']  # EU
        df['Ad Group Name'] = 'Not available'

    if campaign_type == 'SD':
        df['Conversions'] = df['14 Day Total Orders (#)']  # 14 Day Total Units (#)']
        df['Match Type'] = '-'
        try:
            df['Sales'] = df['14 Day Total Sales ']  # US
        except:
            df['Sales'] = df['14 Day Total Sales']  # EU

    df['Ad Group Name'].replace('0', "Not available", inplace=True)
    df['Match Type'].replace('0', "-", inplace=True)
    df['Match Type'].replace(0, "-", inplace=True)
    df['Match Type'].fillna("-", inplace=True)
    # df['Date'] = df['Date'].dt.date
    return df


# Search Querry Report
def read_sqr(path):
    path_file = path + '/raw/Sponsored Products Search term report.xlsx'
    df = pd.read_excel(io=path_file, sheet_name='Sponsored Product Search Term R', header=0, engine="openpyxl")
    return df


# read Active Listing Report - We need it to get the price

def active_listing(path):
    path_file = path + '/raw/Active+Listings+Report.txt'
    df = pd.read_table(path_file)  # , encoding="ANSI")
    return df


# Clustering
def read_clustering(path):
    dfclustered = pd.read_csv(path + "/prediction/dfclustered.csv")
    RF_decoding = pd.read_csv(path + "/prediction/RF_decoding.csv")
    return dfclustered, RF_decoding


def write_clustering(dfclustered, RF_decoding, path):
    dfclustered.to_csv(path + "/prediction/dfclustered.csv")
    RF_decoding.to_csv(path + "/prediction/RF_decoding.csv", index=False)


def read_category_listing_report(path):
    path_file = path + '/raw/Category+Listings+Report.xlsx'
    df = pd.read_excel(io=path_file, sheet_name='Template', header=1, engine="openpyxl")
    return df


def valid_price():
    if today > sales_from_date and today < sale_end_date:
        return df['sale_price']
    else:
        return df['standard_price']


# Settings
def read_target(path):
    path_file = path + '/raw/target.xlsx'
    df = pd.read_excel(path_file, sheet_name='target', header=0, engine="openpyxl", converters={'Campaign Id': str})
    # df['Campaign Id']='\'' + df['Campaign Id']
    return df


def add_campaign_to_target(path, campaign_id, campaign_name):
    print(path)
    print(campaign_id)
    print(campaign_name)
    path_file = path + '/raw/target.xlsx'
    wb_obj = openpyxl.load_workbook(path_file)
    ws_obj = wb_obj['target']
    row = ws_obj.max_row + 1
    print(row)
    ws_obj.cell(row=row, column=1).value = campaign_name
    ws_obj.cell(row=row, column=2).value = campaign_id
    wb_obj.save(path_file)


def read_link(path):
    path_file = path + '/raw/link-manual-manual.xlsx'
    df = pd.read_excel(path_file, header=0, engine="openpyxl",
                       converters={'Ad Group source Id': str, 'Ad Group destination Id': str})
    df['Ad Group source Id'] = '\'' + df['Ad Group source Id'].fillna(0).astype(str)
    df['Ad Group destination Id'] = '\'' + df['Ad Group destination Id'].fillna(0).astype(str)
    return df


# Optimization
def read_kalman_state(path):
    R = np.loadtxt(path + "/prediction/R.txt")
    Q = np.loadtxt(path + "/prediction/Q.txt")
    P = np.loadtxt(path + "/prediction/P.txt")
    H = np.loadtxt(path + "/prediction/H.txt")
    F = np.loadtxt(path + "/prediction/F.txt")
    x = pd.read_csv(path + "/prediction/x.txt", squeeze=True)
    return R, Q, P, H, F, x


def write_kalman_state(R, Q, P, H, F, x, path):
    np.savetxt(path + "/prediction/R.txt", R)
    np.savetxt(path + "/prediction/Q.txt", Q)
    np.savetxt(path + "/prediction/P.txt", P)
    np.savetxt(path + "/prediction/H.txt", H)
    np.savetxt(path + "/prediction/F.txt", F)
    x.to_csv(path + "/prediction/x.csv")


# Prediction by date
def append_state_history(date, X_t, M_t, K_t, P_t, Q_t, R_t, S_t, y_t, path):
    X_t['Date'] = date
    try:
        x_hist = pd.read_csv(path + "/prediction/x_hist.csv", index_col=0)
    except:
        x_hist = X_t
    dfmerge = X_t.join(M_t, lsuffix='_pred', rsuffix='_mes')
    dfmerge = dfmerge.join(K_t)
    dfmerge = dfmerge.join(P_t)
    dfmerge = dfmerge.join(Q_t)
    dfmerge = dfmerge.join(R_t)
    dfmerge = dfmerge.join(S_t)
    dfmerge = dfmerge.join(y_t)
    x_hist = pd.concat([x_hist, dfmerge])
    x_hist.drop_duplicates(inplace=True)
    x_hist.to_csv(path + "/prediction/x_hist.csv")


# Bids
def backup_bid_xls(path):
    oldpath = path + '/raw/bulk.xlsx'
    newpath = path + '/output/bulk_to_upload.xlsx'
    shutil.copyfile(oldpath, newpath)


def open_wb(path, file='/output/bulk_to_upload.xlsx'):
    path_file = path + file
    wb_obj = openpyxl.load_workbook(path_file)
    return wb_obj


def open_ws(wb_obj, campaign_type):
    if campaign_type == 'SP':
        ws_name = 'Sponsored Products Campaigns'
    if campaign_type == 'SB':
        ws_name = 'Sponsored Brands Campaigns'
    if campaign_type == 'SD':
        ws_name = 'Sponsored Display Campaigns'
    ws_obj = wb_obj[ws_name]
    return ws_obj


def read_xls_bid(ws_obj, column, row):
    cell_obj = ws_obj.cell(row=row, column=column)
    return cell_obj.value


def write_xls_bid(wb_obj, ws_obj, row, value, campaign_type):
    if campaign_type == 'SP':
        column = 28
    if campaign_type == 'SB':
        # column=19
        column = 22
    if campaign_type == 'SD':
        # column=20
        column = 26
    ws_obj.cell(row=row, column=column).value = value
    ws_obj.cell(row=row, column=3).value = 'update'


def save_xls_bid(wb_obj, path, filename='/output/bulk_to_upload.xlsx'):
    wb_obj.save(path + filename)
